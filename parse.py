from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from fake_useragent import UserAgent
import requests

def parse():
    UserAgent().chrome # маскируемся под пользователя
    url = 'https://omsk.hh.ru/search/vacancy?text=Python&area=68' # передаём необходимый URL адресс
    page = requests.get(url, headers={'User-Agent': UserAgent().chrome}) # отправляем запрос методом Get на данный адресс
    print(page.status_code) # получаем код запроса
    soup = BeautifulSoup(page.text, "html.parser") # передаём страницу в bs4
    items = soup.findAll('a', class_='serp-item__title') # находим контейнер с нужным классом
    vacancy = [] # создаём пустой список вакансий
    for name in items:
        vacancy.append(name.text) # добавляем вакансии в список
    zapis(vacancy)

def zapis(vacancy):
    file = 'SpisokVacancy.xlsx' # создаём файловую переменную
    try:
        wb = load_workbook(file) # пытаемся загрузить файл
    except FileNotFoundError:
        wb = Workbook() # если файл не найден, создаём новый
    ws = wb.active # создаём переменную, указывающую на лист, с которым будем работать
    for element in vacancy:
        text = element # создаём строковую переменную
        ws.append([text]) # заносим строковую переменную в файл (на указанный лист)
    wb.save(file) # сохраняем изменения
    wb.close() # закрываем файл